import requests
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

# URL base para as planilhas
BASE_URL = "https://www2.aneel.gov.br/aplicacoes/tarifa/arquivo/"

# Lista de nomes dos arquivos
file_names = [
    "SPARTA CN SCEE -Roraima - Versão deliberada.xlsx",
    "SPARTA RTA 2024 ENEL RJ.xlsx",
    "SPARTA CN 2024_RTA2024_Light_V11A.xlsx",
    "SPARTA CN RTA 2024 CPFL Santa Cruz bd25.xlsm",
    "SPARTA_RTA 2024 EMT Final_2.xlsx",
    "SPARTA RTA CPFL Paulista 2024.xlsx",
    "SPARTA_RTA 2024 - EMS (V. Final).xlsx",
    "SPARTA RTA 2024 - Neoenergia Coelba CA - final- bd18.xlsx",
    "SPARTA RTA 2024 ENEL CE.xlsx",
    "SPARTA RTA ESE 2024.xlsx",
    "SPARTA CA_RTA_Neoenergia COSERN_2024_v8.xlsx",
    "SPARTA CA RTA 2024 NEOENERGIA PE_V8.xlsx",
    "SPARTA_RTP 2024 - EQTL AL_Pós CP (V.Final).xlsx",
    "SPARTA RTA 2024_Sulgipe.xlsx",
    "SPARTA CN 2024 - RPT 2024 - AME - final com ajuste MMGD - BD38.xlsx",
    "SPARTA CN RTA 2024 Cemig.xlsx",
    "SPARTA CN 2024 RTA EMR v3.xlsx",
    "SPARTA RTA 2024 COPEL-DIS.xlsx",
    "SPARTA CN RTA 2024 Cocel.xlsx",
    "SPARTA RTA 2024 ETO TUST ciclo anterior.xlsx",
    "SPARTA CA Reajuste_24_Enel SP - v10.xlsx",
    "Sparta_RTA 2024 - ESS (Final).xlsx",
    "Sparta_RTA_2024_UHENPAL_Valida.xlsx",
    "Sparta_RTA_2024_HIDROPAN_Valida.xlsx",
    "SPARTA CN - RTA 2024 - Eletrocar - V6b - Valida - versão STR.xlsx",
    "Sparta_RTA_2024_DEMEI_Valida - versão STR.xlsx",
    "SPARTA CA 2024 EDP ES_V9.xlsx",
    "SPARTA CA RTA 2024-EQTL PA.xlsx",
    "SPARTA RTA 2024 RGE.xlsm",
    "Sparta_RTA_2024_MUX_v5.xlsx",
    "SPARTA CN 2024 Versão 4 0 - subsidio.xlsx",
    "SPARTA CA RTA 2024 Elektro - final.xlsx",
    "SPARTA RTA 2024 EQTL MA.xlsx",
    "SPARTA_RTA 2024 - EPB - V. Final.xlsm",
    "SPARTA CN Celesc 2024_23.xlsx",
    "SPARTA CN 2024_ DCELT v22 final.xlsm",
    "SPARTA CN 2024_ EFLUL v2.xlsx",
    "SPARTA CN RTA 2024 João Cesa ajuste cobertura cusd.xlsx",
    "SPARTA RTA 2024_Cooperaliança.xlsx",
    "SPARTA CN 2024 RTA ELFMS_V3 - final.xlsx",
    "SPARTA CN RTA 2024 NDB.xlsx",
    "SPARTA RTA 2024 EQTL GO.xlsx",
    "SPARTA CA RTA CPFL Piratininga 2024 final.xlsx",
    "SPARTA CA_RTA 2024_EDP SP_18.xlsx",
    "SPARTA CN RTA 2024 CHESP - Final - rev.xlsx",
    "SPARTA CA RTA DMED 2024 v7 - cfurh - IPCA - leo.xlsx",
    "SPARTA RTA 2024 CEEE EQTL_ref_12nov.xlsx",
    "SPARTA Equatorial Piaui RTA 2024 final.xlsx",
    "SPARTA_RTA 2024 - ERO (V. Final).xlsx",
    "SPARTA CN RTA 2024 - EAC_20.xlsx",
    "SPARTA CN 2024 Eqtl Amapá.xlsx"  
    # Adicione outros nomes de arquivo aqui
]

# Construir URLs completas
urls = [BASE_URL + file_name for file_name in file_names]

resultados = []

for url in urls:
    print(f"Processando: {url}")
    response = requests.get(url)
    file = BytesIO(response.content)
    wb = load_workbook(filename=file, data_only=True)
    print("Abas disponíveis:", wb.sheetnames)
    try:
        ws = wb['Resultado']
        dados = {
            'URL': url,
            'B4': ws['B4'].value,
            'B5': ws['B5'].value,
            'D4': ws['D4'].value,
            'D5': ws['D5'].value,
            'K10': ws['K10'].value,
            'L10': ws['L10'].value
        }
        resultados.append(dados)
    except KeyError:
        print(f"Aba 'Resultado' não encontrada em {url}")

# Salvar todos os resultados em um único Excel
df = pd.DataFrame(resultados)
df.to_excel('resultado_extraido_varias_planilhas.xlsx', index=False)